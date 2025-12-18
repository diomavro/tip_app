import sys
import json
from pathlib import Path
import pandas as pd

path = Path(__file__).resolve().parents[1] / 'excel' / 'Psistaria Tips.xlsx'
if not path.exists():
    print(f"ERROR: file not found: {path}")
    sys.exit(1)

xls = pd.ExcelFile(path, engine='openpyxl')
print('SHEETS:', xls.sheet_names)
for sheet in xls.sheet_names:
    try:
        df = pd.read_excel(xls, sheet_name=sheet)
        print('\n--- Sheet:', sheet, ' shape=', df.shape)
        with pd.option_context('display.max_rows', 10, 'display.max_columns', 20):
            print(df.head(10).to_csv(index=False))
    except Exception as e:
        print(f'Failed reading sheet {sheet}:', e)

# try to find cells with formulas using openpyxl
try:
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=False)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        formulas = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formulas.append((cell.coordinate, cell.value))
        if formulas:
            print('\nFormulas in sheet', sheet)
            for coord, f in formulas[:20]:
                print(coord, f)
except Exception as e:
    print('openpyxl formulas scan failed:', e)
